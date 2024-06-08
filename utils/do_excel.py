from openpyxl import load_workbook

class DoExcel:
    def __init__(self, filename):
        self.wb = load_workbook(filename)
        self.ws = self.wb.active

    def get_data_from_excel(self):
        data = []
        if self.ws.max_row > 1:  # 确保工作表至少有一行数据
            headers = [cell.value for cell in self.ws[1]]  # 读取第一行作为列标题
            for row in self.ws.iter_rows(min_row=2, values_only=True):
                data.append(dict(zip(headers, row)))  # 使用列标题作为键
        return data